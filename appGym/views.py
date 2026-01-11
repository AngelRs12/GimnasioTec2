from datetime import date
import time
from django.shortcuts import render, redirect
from django.db import connection
from django.contrib import messages
from django.contrib.auth.hashers import make_password, check_password
from django.http import Http404, HttpResponseForbidden, JsonResponse
from django.db import IntegrityError, transaction
from django.views.decorators.http import require_POST
import io
import base64
from django.http import HttpResponse
import os
from django.conf import settings
from django.contrib.auth.hashers import make_password
import uuid
from django.http import JsonResponse
from django.db import connection
import json
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import connection
import io

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
from django.http import HttpResponse


def index(request):
    return render(request, 'gym/index.html')

def usuarios(request):
    return render(request, 'gym/usuarios.html')

def entradas_salidas(request):
    return render(request, 'gym/entradas_salidas.html')

def graficas(request):
    return render(request, 'gym/graficas.html')

def membresias(request):
    return render(request, 'gym/membresias.html')

def observaciones(request):
    return render(request, 'gym/observaciones.html')

def reglamento(request):
    return render(request, "gym/reglamento.html")

def horario(request):
    return render(request, 'gym/horario.html')

def reportes_exportacion(request):
    return render(request, "gym/reportes_exportacion.html")

def entrenadores(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM obtener_entrenadores()")
        rows = cursor.fetchall()

    entrenadores_list = []
    for r in rows:
        entrenadores_list.append({
            "id_entrenador": r[0],
            "Nombres": r[1],
            "ApellidoM": r[2],
            "ApellidoP": r[3],
            "descrpicion": r[4],
            "url_img": r[5]
        })
    print(entrenadores_list)
    return render(request, "gym/entrenadores.html", {
        "entrenadores": entrenadores_list,
        "MEDIA_URL": settings.MEDIA_URL
    })

def actividades(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM sp_select_actividades()")
        columns = [col[0] for col in cursor.description]
        actividades = [dict(zip(columns, row)) for row in cursor.fetchall()]
    return render(request, "gym/actividades.html", {"actividades": actividades})
    
def acercade(request):
    return render(request, 'gym/acercade.html')

def sesion(request):
    return render(request, 'gym/sesion.html')

def login(request):
    if request.method == "POST":
        usuario = request.POST.get("usuario")
        password = request.POST.get("password")

        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT "Password" FROM public."Usuarios_admin" WHERE "Usuario" = %s;',
                [usuario]
            )
            fila = cursor.fetchone()

        if not fila:
            return render(request, 'gym/sesion.html', {
                "error": "Usuario o contraseña incorrectos"
            })

        password_hash = fila[0]

        if check_password(password, password_hash):
            request.session["usuario_admin"] = usuario
            return redirect("index")

        return render(request, 'gym/sesion.html', {
            "error": "Usuario o contraseña incorrectos"
        })

def logout(request):
    request.session.flush()
    return redirect("index") 

def administradores(request):   
    return render(request, 'gym/administradores.html')

def crear_admin(request):
    if request.method == "POST":
        usuario = (request.POST.get("usuario") or "").strip()
        password = (request.POST.get("password") or "").strip()

        if not usuario or not password:
            messages.error(request, "Por favor complete todos los campos.")
            return render(request, "gym/administradores.html")

        try:
            with transaction.atomic():
                password_hash = make_password(password)
                with connection.cursor() as cursor:
                    cursor.execute(
                        'INSERT INTO public."Usuarios_admin" ("Usuario", "Password") VALUES (%s, %s);',
                        [usuario, password_hash]
                    )

            messages.success(request, "Administrador creado correctamente.")
        except IntegrityError as e:
            messages.error(request, "No fue posible crear el administrador: el nombre de usuario ya existe.")
        except Exception as e:
            messages.error(request, f"No fue posible crear el administrador: {str(e)}")
        return render(request, "gym/administradores.html")
    return render(request, "gym/administradores.html")

def buscar_admin(request):
    # Obtener parámetros
    id_param = request.GET.get("id_admin") or request.GET.get("id") or ""
    id_param = id_param.strip()
    nombre = (request.GET.get("usuario") or request.GET.get("nombre") or "").strip()
    
    if not id_param and not nombre: 
        return render(request, "gym/administradores.html")
    try:
        rows = []
        with connection.cursor() as cursor:
            if nombre:
                cursor.execute(
                    """
                    SELECT
                      "id_admin",
                      "Usuario" AS usuario
                    FROM public."Usuarios_admin"
                    WHERE "Usuario" ILIKE %s
                    ORDER BY "Usuario"
                    LIMIT 200;
                    """,
                    [f"%{nombre}%"]
                )
                rows = cursor.fetchall()
            else:
                try:
                    id_int = int(id_param)
                except (ValueError, TypeError):
                    rows = []
                else:
                    cursor.execute(
                        """
                        SELECT "id_admin", "Usuario"
                        FROM public."Usuarios_admin"
                        WHERE "id_admin" = %s
                        LIMIT 1;
                        """,
                        [id_int]
                    )
                    rows = cursor.fetchall()
        resultados = [{"id_admin": r[0], "usuario": r[1]} for r in rows]
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True, "resultados": resultados})
        else:
            return render(request, "gym/administradores.html", {
                "resultados": resultados,
                "query_usuario": nombre,
                "query_id": id_param
            })

    except Exception as e:
        err = str(e)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "error": err}, status=500)
        return render(request, "gym/administradores.html", {"error": f"No fue posible realizar la búsqueda: {err}"})

def eliminar_admin(request):
    if request.method != "POST":
        return redirect("administradores")

    if "usuario_admin" not in request.session:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
        messages.error(request, "Debes iniciar sesión como administrador.")
        return redirect("sesion")

    usuario = (request.POST.get("usuario") or "").strip()
    id_param = request.POST.get("id_admin") or request.POST.get("id")

    try:
        with connection.cursor() as cursor:
            if id_param and not usuario:
                try:
                    id_int = int(id_param)
                except ValueError:
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                        return JsonResponse({"success": False, "error": "ID inválido."}, status=400)
                    return render(request, "gym/administradores.html", {"error": "ID inválido."})

                cursor.execute(
                    """
                    SELECT usuario FROM (
                      SELECT row_number() OVER (ORDER BY "Usuario") AS id_admin,
                             "Usuario" AS usuario
                      FROM public."Usuarios_admin" ORDER BY "Usuario"
                    ) t WHERE t.id_admin = %s LIMIT 1;
                    """,
                    [id_int]
                )
                fila = cursor.fetchone()
                if not fila:
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                        return JsonResponse({"success": False, "error": "Administrador no encontrado."}, status=404)
                    return render(request, "gym/administradores.html", {"error": "Administrador no encontrado."})
                usuario = fila[0]

            if not usuario:
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Usuario a eliminar no proporcionado."}, status=400)
                return render(request, "gym/administradores.html", {"error": "Usuario a eliminar no proporcionado."})

            cursor.execute(
                'DELETE FROM public."Usuarios_admin" WHERE "Usuario" = %s;',
                [usuario]
            )
            if cursor.rowcount == 0:
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Administrador no encontrado."}, status=404)
                return render(request, "gym/administradores.html", {"error": "Administrador no encontrado."})

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True, "mensaje": f"Administrador '{usuario}' eliminado correctamente."})
        messages.success(request, f"Administrador '{usuario}' eliminado correctamente.")
        return redirect("administradores")

    except Exception as e:
        err = str(e)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "error": err}, status=500)
        return render(request, "gym/administradores.html", {"error": f"No fue posible eliminar el administrador: {err}"})

def editar_admin(request):
    if request.method != "POST":
        return redirect("administradores")

    if "usuario_admin" not in request.session:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "error": "No autorizado"}, status=403)
        messages.error(request, "Debes iniciar sesión como administrador.")
        return redirect("sesion")

    usuario = (request.POST.get("usuario") or "").strip()
    password = (request.POST.get("password") or "").strip()
    id_param = request.POST.get("id_admin") or request.POST.get("id")

    if not password:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "error": "La nueva contraseña es obligatoria."}, status=400)
        return render(request, "gym/administradores.html", {"error": "La nueva contraseña es obligatoria."})

    try:
        with connection.cursor() as cursor:
            if id_param and not usuario:
                try:
                    id_int = int(id_param)
                except ValueError:
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                        return JsonResponse({"success": False, "error": "ID inválido."}, status=400)
                    return render(request, "gym/administradores.html", {"error": "ID inválido."})

                cursor.execute(
                    """
                    SELECT usuario FROM (
                      SELECT row_number() OVER (ORDER BY "Usuario") AS id_admin,
                             "Usuario" AS usuario
                      FROM public."Usuarios_admin" ORDER BY "Usuario"
                    ) t WHERE t.id_admin = %s LIMIT 1;
                    """,
                    [id_int]
                )
                fila = cursor.fetchone()
                if not fila:
                    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                        return JsonResponse({"success": False, "error": "Administrador no encontrado."}, status=404)
                    return render(request, "gym/administradores.html", {"error": "Administrador no encontrado."})
                usuario = fila[0]

            if not usuario:
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Usuario a editar no proporcionado."}, status=400)
                return render(request, "gym/administradores.html", {"error": "Usuario a editar no proporcionado."})

            password_hash = make_password(password)
            cursor.execute(
                'UPDATE public."Usuarios_admin" SET "Password" = %s WHERE "Usuario" = %s;',
                [password_hash, usuario]
            )
            if cursor.rowcount == 0:
                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse({"success": False, "error": "Administrador no encontrado."}, status=404)
                return render(request, "gym/administradores.html", {"error": "Administrador no encontrado."})

        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": True, "mensaje": f"Contraseña de '{usuario}' actualizada correctamente."})
        messages.success(request, f"Contraseña de '{usuario}' actualizada correctamente.")
        return redirect("administradores")

    except Exception as e:
        err = str(e)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return JsonResponse({"success": False, "error": err}, status=500)
        return render(request, "gym/administradores.html", {"error": f"No fue posible editar el administrador: {err}"})
    
def guardar_foto_usuario(file):
    if not file:
        return None

    carpeta = os.path.join(settings.MEDIA_ROOT, "fotos_usuarios")
    os.makedirs(carpeta, exist_ok=True)

    extension = os.path.splitext(file.name)[1]
    nombre_img = f"{uuid.uuid4().hex}{extension}"

    ruta_absoluta = os.path.join(carpeta, nombre_img)

    with open(ruta_absoluta, "wb+") as destino:
        for chunk in file.chunks():
            destino.write(chunk)

    return f"fotos_usuarios/{nombre_img}"

def borrar_foto_usuario(ruta_foto):
    if not ruta_foto:
        return
    ruta_absoluta = os.path.join(settings.MEDIA_ROOT, ruta_foto)
    if os.path.exists(ruta_absoluta):
        try:
            os.remove(ruta_absoluta)
        except Exception:
            pass
        
def gestion_usuarios(request):
    if request.method == "POST":
        accion = request.POST.get("accion")

        nombre = request.POST.get("nombre")
        apellido_paterno = request.POST.get("apellido_paterno")
        apellido_materno = request.POST.get("apellido_materno")
        tipo_usuario = request.POST.get("tipo_usuario")

        no_control = request.POST.get("no_control")
        equipo = request.POST.get("equipo")
        numero_empleado = request.POST.get("numero_empleado")

        foto = request.FILES.get("foto")
        if accion == "agregar":
            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT insertar_usuario_general(%s, %s, %s, %s, %s, %s, %s)
                    """, [
                        nombre,
                        apellido_paterno,
                        apellido_materno,
                        tipo_usuario,
                        no_control,
                        numero_empleado,
                        equipo
                    ])

                    nuevo_id = cursor.fetchone()[0]

                    ruta_foto = guardar_foto_usuario(foto)
                    if ruta_foto:
                        cursor.execute("""
                            UPDATE usuario
                            SET foto = %s
                            WHERE id_usuario = %s
                        """, [ruta_foto, nuevo_id])

                if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                    return JsonResponse({
                        "success": True,
                        "mensaje": f"Usuario agregado correctamente (ID: {nuevo_id})."
                    })

                return render(
                    request,
                    "gym/usuarios.html",
                    {"mensaje": f"Usuario agregado correctamente (ID: {nuevo_id})."}
                )

            except Exception as e:
                error_msg = str(e).split("CONTEXT:")[0].strip()

                return JsonResponse(
                    {"success": False, "error": error_msg},
                    status=400
                )

        if accion == "editar":
            id_usuario = request.POST.get("id_usuario")

            try:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        SELECT foto
                        FROM usuario
                        WHERE id_usuario = %s
                    """, [id_usuario])

                    fila = cursor.fetchone()
                    foto_actual = fila[0] if fila else None
                    cursor.execute("""
                        SELECT editar_usuario_general(%s, %s, %s, %s, %s, %s, %s, %s)
                    """, [
                        id_usuario,
                        nombre,
                        apellido_paterno,
                        apellido_materno,
                        tipo_usuario,
                        no_control if no_control else None,
                        numero_empleado if numero_empleado else None,
                        equipo if equipo else None
                    ])
                    mensaje = cursor.fetchone()[0]
                    ruta_foto_nueva = None

                    if foto:
                        ruta_foto_nueva = guardar_foto_usuario(foto)

                        if ruta_foto_nueva:
                            cursor.execute("""
                                UPDATE usuario
                                SET foto = %s
                                WHERE id_usuario = %s
                            """, [ruta_foto_nueva, id_usuario])

                            borrar_foto_usuario(foto_actual)

                return JsonResponse({
                    "success": True,
                    "mensaje": mensaje,
                    "id_usuario": id_usuario,
                    "foto": f"{settings.MEDIA_URL}{ruta_foto_nueva}" if ruta_foto_nueva else None
                })

            except Exception as e:
                error_msg = str(e).split("CONTEXT:")[0].strip()
                return JsonResponse(
                    {"success": False, "error": error_msg},
                    status=400
                )

    elif request.method == "GET" and request.headers.get("X-Requested-With") == "XMLHttpRequest":
        id_usuario = request.GET.get("id_usuario", "").strip()
        nombre = request.GET.get("nombre", "").strip()

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT * FROM buscar_usuario(%s, %s)",
                    [id_usuario, nombre]
                )

                columnas = [col[0] for col in cursor.description]
                resultados = [
                    dict(zip(columnas, fila))
                    for fila in cursor.fetchall()
                ]

            return JsonResponse({"success": True, "resultados": resultados})

        except Exception as e:
            error_str = str(e).split("CONTEXT:")[0].strip()
            return JsonResponse(
                {"success": False, "error": error_str},
                status=400
            )

    return render(request, "gym/usuarios.html")

def registrar_ingreso(request):
    if request.method == "POST":
        id_usuario = request.POST.get("id_usuario")
        tipo = request.POST.get("tipo")

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT insertar_ingreso(%s, %s);",
                    [id_usuario, tipo]
                )
                mensaje = cursor.fetchone()[0]

            if mensaje.startswith("Registro inválido"):
                return JsonResponse({
                    "success": False,
                    "error": mensaje
                })

            return JsonResponse({
                "success": True,
                "mensaje": mensaje
            })

        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            })

    return JsonResponse({
        "success": False,
        "error": "Método no permitido"
    })

def guardar_observacion(request):
    if request.method == "POST":
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion")
        fecha = request.POST.get("fecha_publicacion")

        if not titulo or not descripcion or not fecha:
            return JsonResponse({"success": False, "error": "Todos los campos son obligatorios."})

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT insertar_observacion(%s, %s, %s);",
                    [titulo, descripcion, fecha]
                )
                mensaje = cursor.fetchone()[0]

            return JsonResponse({"success": True, "mensaje": mensaje})

        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método no permitido"})


def listar_observaciones(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM listar_observaciones();")
        rows = cursor.fetchall()

    columnas = ["id_observacion", "titulo", "descripcion", "fecha_publicacion"]
    data = [dict(zip(columnas, fila)) for fila in rows]

    return JsonResponse(data, safe=False)


def editar_observacion_view(request):
    if request.method == "POST":
        id_observacion = request.POST.get("id_observacion")
        descripcion = request.POST.get("descripcion")

        if not id_observacion or not descripcion:
            return JsonResponse({"success": False, "error": "ID y descripción son obligatorios."})

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT editar_observacion(%s, %s);",
                    [id_observacion, descripcion]
                )
                mensaje = cursor.fetchone()[0]

            return JsonResponse({"success": True, "mensaje": mensaje})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método no permitido"})

def eliminar_observacion_view(request):
    if request.method == "POST":
        id_observacion = request.POST.get("id_observacion")

        if not id_observacion:
            return JsonResponse({"success": False, "error": "ID es obligatorio."})

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT eliminar_observacion(%s);",
                    [id_observacion]
                )
                mensaje = cursor.fetchone()[0]

            return JsonResponse({"success": True, "mensaje": mensaje})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método no permitido"})

def eliminar_usuario(request):
    if request.method == "POST" and request.headers.get("x-requested-with") == "XMLHttpRequest":
        id_usuario = request.POST.get("id_usuario")

        if not id_usuario:
            return JsonResponse({
                "success": False,
                "error": "ID de usuario no proporcionado."
            })

        try:
            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT foto
                    FROM usuario
                    WHERE id_usuario = %s
                """, [id_usuario])

                fila = cursor.fetchone()
                foto = fila[0] if fila else None

                cursor.execute(
                    "SELECT public.eliminar_usuario(%s);",
                    [int(id_usuario)]
                )

            if foto:
                ruta_fisica = os.path.join(settings.MEDIA_ROOT, foto)
                if os.path.exists(ruta_fisica):
                    os.remove(ruta_fisica)

            return JsonResponse({
                "success": True,
                "mensaje": f"Usuario {id_usuario} eliminado correctamente."
            })

        except Exception as e:
            error_msg = str(e).split("CONTEXT:")[0].strip()
            return JsonResponse({
                "success": False,
                "error": error_msg
            }, status=400)

    return JsonResponse({
        "success": False,
        "error": "Método no permitido."
    }, status=405)
def crear_membresia(request):
    if request.method != "POST":
        return JsonResponse({
            "status": "error",
            "message": "Método inválido, use POST"
        }, status=400)

    nombre = request.POST.get("nombre_tipo")
    duracion = request.POST.get("duracion")
    costo = request.POST.get("costo_tipo")

    if not nombre or not duracion or not costo:
        return JsonResponse({
            "status": "error",
            "message": "Faltan campos obligatorios"
        }, status=400)

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT insertar_membresia_general(%s, %s, %s);",
                [nombre, duracion, costo]
            )

        return JsonResponse({
            "status": "success",
            "message": f"Membresía '{nombre}' creada correctamente"
        }, status=200)

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": str(e)
        }, status=400)
        
def obtener_membresias(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM obtener_membresias_general();")
        rows = cursor.fetchall()

    membresias = []
    for row in rows:
        membresias.append({
            "id": row[0],
            "nombre": row[1],
            "duracion": row[2],
            "costo": row[3],
        })

    return JsonResponse({
        "status": "ok",
        "data": membresias
    })
    
@require_POST    
def editar_membresia(request):
    nombre = request.POST.get("nombre")
    duracion = request.POST.get("duracion")
    costo = request.POST.get("costo")

    if not all([nombre, duracion, costo]):
        return JsonResponse({
            "status": "error",
            "message": "Faltan parámetros."
        })

    with connection.cursor() as cursor:
        cursor.execute(
            "SELECT editar_membresia_por_nombre(%s, %s, %s);",
            [nombre, duracion, costo]
        )
        result = cursor.fetchone()[0]

    return JsonResponse(result)

@require_POST
def eliminar_membresia(request):
    nombre = request.POST.get("nombre")
    print(request.POST)

    if not nombre:
        return JsonResponse({"status": "error", "message": "Nombre requerido."})

    with connection.cursor() as cursor:
        cursor.execute("SELECT eliminar_membresia_por_nombre(%s);", [nombre])
        result = cursor.fetchone()[0]  # El JSON devuelto por la función

    return JsonResponse(result)


from django.http import JsonResponse
from django.db import connection


def uso_gimnasio_data(request):
    query = """
        SELECT
          EXTRACT(DOW FROM fecha) AS dia,
          COUNT(*) AS total
        FROM ingresos
        WHERE tipo = 'ENTRADA'
          AND DATE_TRUNC('week', fecha) = DATE_TRUNC('week', CURRENT_DATE)
          AND EXTRACT(DOW FROM fecha) BETWEEN 1 AND 5
        GROUP BY dia
        ORDER BY dia;
    """

    try:
        with connection.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    dias_map = {
        1: "Lunes",
        2: "Martes",
        3: "Miércoles",
        4: "Jueves",
        5: "Viernes"
    }

    labels = []
    data = []

    # Inicializar conteos en 0 para toda la semana
    conteo_por_dia = {dia: 0 for dia in range(1, 6)}

    for dia, total in rows:
        conteo_por_dia[int(dia)] = int(total)

    for dia in range(1, 6):
        labels.append(dias_map[dia])
        data.append(conteo_por_dia[dia])

    return JsonResponse({
        "labels": labels,
        "data": data
    })


def uso_gimnasio_por_hora_data(request):
    query = """
        SELECT 
          EXTRACT(DOW FROM fecha) AS dia,
          EXTRACT(HOUR FROM fecha) AS hora,
          COUNT(*) AS total
        FROM ingresos
        WHERE tipo = 'ENTRADA'
          AND EXTRACT(DOW FROM fecha) BETWEEN 1 AND 5
          AND EXTRACT(HOUR FROM fecha) BETWEEN 8 AND 23
        GROUP BY dia, hora
        ORDER BY dia, hora;
    """

    try:
        with connection.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

    dias_map = {
        1: "Lunes",
        2: "Martes",
        3: "Miércoles",
        4: "Jueves",
        5: "Viernes"
    }

    conteo = {}
    for d in range(1, 6):
        for h in range(8, 24):
            conteo[(d, h)] = 0

    for dia, hora, total in rows:
        conteo[(int(dia), int(hora))] = int(total)

    labels = []
    data = []

    for d in range(1, 6):
        for h in range(8, 24):
            labels.append(f"{dias_map[d]} {h:02d}:00")
            data.append(conteo[(d, h)])

    return JsonResponse({
        "labels": labels,
        "data": data
    })

def _fetch_users_rows():
    query = """
       SELECT u.id_usuario,
              u.nombres,
              u.apellido_paterno,
              u.apellido_materno,
              COALESCE(v.tipo, 'desconocido') AS tipo,
              CASE
                  WHEN ma.id_usuario IS NOT NULL THEN 'con_membresia'
                  ELSE 'sin_membresia'
              END AS membresia
       FROM public.usuario u
       LEFT JOIN public.vista_tipo_usuario v ON v.id_usuario = u.id_usuario
       LEFT JOIN public.vista_membresia_activa ma ON ma.id_usuario = u.id_usuario
       ORDER BY u.id_usuario;
    """
    with connection.cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()
    return rows


def _build_counts_from_rows(rows):
    counts = {}
    for r in rows:
        tipo = (r[4] or 'desconocido')
        counts[tipo] = counts.get(tipo, 0) + 1

    expected_order = ['alumno', 'representativo', 'externo', 'con_membresia']
    friendly = {'con_membresia': 'Con membresía activa',
    'representativo': 'Representativo',
    'empleado': 'Empleado',
    'alumno': 'Alumno',
    'externo': 'Externo',
    'desconocido': 'Desconocido'}

    labels = []
    data = []
    for k in expected_order:
        labels.append(friendly.get(k, k.capitalize()))
        data.append(int(counts.get(k, 0)))

    others = [k for k in counts.keys() if k not in expected_order]
    for k in sorted(others):
        labels.append(friendly.get(k, k.capitalize()))
        data.append(int(counts.get(k, 0)))

    resumen = [{"tipo": labels[i], "conteo": data[i]} for i in range(len(labels))]
    total = sum(data)
    return labels, data, resumen, total

def reportes_view(request):
    try:
        rows = _fetch_users_rows()
        labels, data, resumen, total = _build_counts_from_rows(rows)
    except Exception as e:
        return render(request, "gym/graficas.html", {"error": f"Error al consultar la BD: {e}"})

    context = {
        "labels_json": json.dumps(labels),
        "data_json": json.dumps(data),
        "total": total,
        "resumen": resumen,
    }
    return render(request, "gym/graficas.html", context)

def reportes_data(request):
    try:
        rows = _fetch_users_rows()
        labels, data, resumen, total = _build_counts_from_rows(rows)
        return JsonResponse({"labels": labels, "data": data, "total": total})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import io
from django.db import connection


def reporte_usuarios_excel(request):
    wb = Workbook()
    wb.remove(wb.active)

    # =========================
    # HOJAS DE USUARIOS
    # =========================
    consultas = [
        (
            "Alumnos",
            """
            SELECT u.*, a.*
            FROM public.alumno a
            JOIN public.usuario u ON u.id_usuario = a.id_usuario;
            """
        ),
        (
            "Empleados",
            """
            SELECT u.*, e.*
            FROM public.empleado e
            JOIN public.usuario u ON u.id_usuario = e.id_usuario;
            """
        ),
        (
            "Externos",
            """
            SELECT u.*, ex.*
            FROM public.externo ex
            JOIN public.usuario u ON u.id_usuario = ex.id_usuario;
            """
        ),
        (
            "Representativos",
            """
            SELECT u.*, r.*
            FROM public.representativos r
            JOIN public.usuario u ON u.id_usuario = r.id_usuario;
            """
        ),
    ]

    for nombre_hoja, query in consultas:
        with connection.cursor() as cursor:
            cursor.execute(query)
            rows = cursor.fetchall()
            headers = [col[0] for col in cursor.description]

        ws = wb.create_sheet(title=nombre_hoja)
        ws.append(headers)

        for row in rows:
            ws.append(row)

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22

    # =========================
    # HOJA DE DATOS PARA GRÁFICA
    # =========================
    ws_data = wb.create_sheet(title="Resumen Grafica")

    ws_data.append(["Tipo de usuario", "Cantidad"])

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT tipo, COUNT(*)
            FROM vista_tipo_usuario
            WHERE tipo <> 'desconocido'
            GROUP BY tipo
            ORDER BY tipo;
        """)
        resultados = cursor.fetchall()

    for tipo, total in resultados:
        ws_data.append([tipo.replace('_', ' ').title(), total])

    for col in range(1, 3):
        ws_data.column_dimensions[get_column_letter(col)].width = 30

    # =========================
    # HOJA DE GRÁFICA
    # =========================
    ws_chart = wb.create_sheet(title="Grafica Usuarios")

    chart = BarChart()
    chart.title = "Usuarios por tipo"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Tipo de usuario"
    chart.style = 10

    data = Reference(
        ws_data,
        min_col=2,
        min_row=1,
        max_row=ws_data.max_row
    )

    categories = Reference(
        ws_data,
        min_col=1,
        min_row=2,
        max_row=ws_data.max_row
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.height = 12
    chart.width = 22

    ws_chart.add_chart(chart, "B2")

    # =========================
    # EXPORTAR EXCEL
    # =========================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="usuarios.xlsx"'
    return response

def reporte_ingresos_excel(request):
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja vacía inicial
    
    dias = [
        ("Lunes", 1),
        ("Martes", 2),
        ("Miércoles", 3),
        ("Jueves", 4),
        ("Viernes", 5),
    ]

    base_query = """
        SELECT
            u.id_usuario,
            u.nombres,
            u.apellido_paterno,
            u.apellido_materno,
            EXTRACT(HOUR FROM i.fecha) AS hora,
            i.fecha AS fecha_completa
        FROM ingresos i
        JOIN public.usuario u ON u.id_usuario = i.id_usuario
        WHERE i.tipo = 'ENTRADA'
          AND EXTRACT(DOW FROM i.fecha) = %s
          AND EXTRACT(HOUR FROM i.fecha) BETWEEN 8 AND 23
          AND (%s IS NULL OR i.fecha >= %s)
          AND (%s IS NULL OR i.fecha <= %s)
        ORDER BY
            hora,
            i.fecha;
    """

    # =========================
    # HOJAS POR DÍA (DETALLE)
    # =========================
    for nombre_dia, dow in dias:
        with connection.cursor() as cursor:
            cursor.execute(base_query, [dow,inicio, inicio, fin, fin])
            rows = cursor.fetchall()
            headers = [col[0] for col in cursor.description]

        ws = wb.create_sheet(title=nombre_dia)
        ws.append(headers)

        for row in rows:
            ws.append(row)

        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22

    # =========================
    # HOJA RESUMEN (PARA GRÁFICA)
    # =========================
    ws_resumen = wb.create_sheet(title="Resumen Ingresos")
    ws_resumen.append(["Día", "Total de ingresos"])

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                CASE EXTRACT(DOW FROM fecha)
                    WHEN 1 THEN 'Lunes'
                    WHEN 2 THEN 'Martes'
                    WHEN 3 THEN 'Miércoles'
                    WHEN 4 THEN 'Jueves'
                    WHEN 5 THEN 'Viernes'
                END AS dia,
                COUNT(*) AS total
            FROM ingresos
            WHERE tipo = 'ENTRADA'
              AND EXTRACT(DOW FROM fecha) BETWEEN 1 AND 5
              AND EXTRACT(HOUR FROM fecha) BETWEEN 8 AND 23
            GROUP BY dia
            ORDER BY
                MIN(EXTRACT(DOW FROM fecha));
        """)
        resumen_rows = cursor.fetchall()

    for dia, total in resumen_rows:
        ws_resumen.append([dia, total])

    for i in range(1, 3):
        ws_resumen.column_dimensions[get_column_letter(i)].width = 30

    # =========================
    # HOJA DE GRÁFICA
    # =========================
    ws_chart = wb.create_sheet(title="Grafica Ingresos")

    chart = BarChart()
    chart.title = "Ingresos al gimnasio por día"
    chart.y_axis.title = "Cantidad de ingresos"
    chart.x_axis.title = "Día"
    chart.style = 10

    data = Reference(
        ws_resumen,
        min_col=2,
        min_row=1,
        max_row=ws_resumen.max_row
    )

    categories = Reference(
        ws_resumen,
        min_col=1,
        min_row=2,
        max_row=ws_resumen.max_row
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.width = 22
    chart.height = 12

    ws_chart.add_chart(chart, "B2")

    # =========================
    # EXPORTAR EXCEL
    # =========================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="ingresos_por_dia.xlsx"'
    )
    return response


def reporte_membresias_excel(request):
    wb = Workbook()
    wb.remove(wb.active)

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    params = [inicio, inicio, fin, fin]
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    params = [inicio, inicio, fin, fin]
    query = """
        SELECT
            u.id_usuario,
            u.nombres,
            m.no_membresia,
            m.fecha_inicial,
            m.fecha_final,
            m.status
        FROM public.membresias m
        JOIN public.usuario u ON u.id_usuario = m.id_usuario
        WHERE (%s IS NULL OR m.fecha_inicial >= %s)
          AND (%s IS NULL OR m.fecha_final <= %s)
        ORDER BY m.fecha_inicial DESC
    """

    with connection.cursor() as cursor:
        if "%s" in query:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        rows = cursor.fetchall()
        headers = [col[0] for col in cursor.description]

    ws_detalle = wb.create_sheet(title="Membresias")
    ws_detalle.append(headers)

    for row in rows:
        ws_detalle.append(row)

    for i in range(1, len(headers) + 1):
        ws_detalle.column_dimensions[get_column_letter(i)].width = 22

    # =========================
    # HOJA RESUMEN (PARA GRAFICA)
    # =========================
    ws_resumen = wb.create_sheet(title="Resumen Membresias")
    ws_resumen.append(["Status", "Cantidad"])

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                status,
                COUNT(*) AS total
            FROM membresias
            GROUP BY status
            ORDER BY status;
        """)
        resumen_rows = cursor.fetchall()

    for status, total in resumen_rows:
        ws_resumen.append([status.title(), total])

    for i in range(1, 3):
        ws_resumen.column_dimensions[get_column_letter(i)].width = 30

    # =========================
    # HOJA DE GRAFICA
    # =========================
    ws_chart = wb.create_sheet(title="Grafica Membresias")

    chart = BarChart()
    chart.title = "Membresías por estado"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Estado"
    chart.style = 10

    data = Reference(
        ws_resumen,
        min_col=2,
        min_row=1,
        max_row=ws_resumen.max_row
    )

    categories = Reference(
        ws_resumen,
        min_col=1,
        min_row=2,
        max_row=ws_resumen.max_row
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.width = 22
    chart.height = 12

    ws_chart.add_chart(chart, "B2")

    # =========================
    # EXPORTAR EXCEL
    # =========================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="membresias.xlsx"'
    )
    return response


def reporte_observaciones_excel(request):
    wb = Workbook()
    wb.remove(wb.active)
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    params = [inicio, inicio, fin, fin]
    # =========================
    # HOJA DETALLE OBSERVACIONES
    # =========================
    query = """
        SELECT
            id_observacion,
            fecha_observacion,
            titulo,
            descripcion
        FROM public.observaciones
        WHERE (%s IS NULL OR fecha_observacion >= %s)
          AND (%s IS NULL OR fecha_observacion <= %s)
        ORDER BY fecha_observacion DESC
    """

    with connection.cursor() as cursor:
        if "%s" in query:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        
        rows = cursor.fetchall()
        headers = [col[0] for col in cursor.description]

    ws_detalle = wb.create_sheet(title="Observaciones")
    ws_detalle.append(headers)

    for row in rows:
        ws_detalle.append(row)

    for i in range(1, len(headers) + 1):
        ws_detalle.column_dimensions[get_column_letter(i)].width = 22

    # =========================
    # HOJA RESUMEN (PARA GRAFICA)
    # =========================
    ws_resumen = wb.create_sheet(title="Resumen Observaciones")
    ws_resumen.append(["Mes", "Total de observaciones"])

    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT
                TO_CHAR(fecha_observacion, 'YYYY-MM') AS mes,
                COUNT(*) AS total
            FROM observaciones
            GROUP BY mes
            ORDER BY mes;
        """)
        resumen_rows = cursor.fetchall()

    for mes, total in resumen_rows:
        ws_resumen.append([mes, total])

    for i in range(1, 3):
        ws_resumen.column_dimensions[get_column_letter(i)].width = 30

    # =========================
    # HOJA DE GRAFICA
    # =========================
    ws_chart = wb.create_sheet(title="Grafica Observaciones")

    chart = BarChart()
    chart.title = "Observaciones por mes"
    chart.y_axis.title = "Cantidad"
    chart.x_axis.title = "Mes"
    chart.style = 10

    data = Reference(
        ws_resumen,
        min_col=2,
        min_row=1,
        max_row=ws_resumen.max_row
    )

    categories = Reference(
        ws_resumen,
        min_col=1,
        min_row=2,
        max_row=ws_resumen.max_row
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart.width = 22
    chart.height = 12

    ws_chart.add_chart(chart, "B2")

    # =========================
    # EXPORTAR EXCEL
    # =========================
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="observaciones.xlsx"'
    )
    return response

def _exportar_excel(nombre_archivo, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append(headers)

    for row in rows:
        ws.append(row)

    for i, col in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


def actividad_eliminar(request, id_actividad):
    if request.method != "POST":
        return JsonResponse({'ok': False, 'msg': "Método inválido"}, status=400)

    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT sp_delete_actividad(%s);', [id_actividad])

        return JsonResponse({'ok': True, 'message':'Actividad eliminada correctamente', 'success':True})

    except Exception as e:
        return JsonResponse({'ok': False, 'message': str(e)}, status=500)

def actividad_editar(request, id_actividad):
    if request.method != "POST":
        return JsonResponse({'ok': False, 'msg': "Método no permitido"}, status=405)

    nombre = request.POST.get('nombre')
    descripcion = request.POST.get('descripcion')
    horario = request.POST.get('horario')

    if not nombre:
        return JsonResponse({'ok': False, 'msg': "El nombre es obligatorio"}, status=422)

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT sp_update_actividad(%s, %s, %s, %s);",
                [id_actividad, nombre, descripcion, horario]
            )
        return JsonResponse({'ok': True, 'message':'Actividad editada correctamente', 'success':True})

    except Exception as e:
        return JsonResponse({'ok': False, 'message': str(e)}, status=500)
    
def actividad_agregar(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        descripcion = request.POST.get("descripcion")
        horario = request.POST.get("horario")

    if not nombre:
        return JsonResponse({'ok': False, 'msg': "Falta nombre"}, status=422)

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                'SELECT sp_insert_actividad(%s, %s, %s);',
                [nombre, descripcion, horario]
            )

        return JsonResponse({
            "success": True,
            "message": "Actividad agregada correctamente."
        })

    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Error al agregar la actividad: {e}"
            })
        
def actividades_json(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM sp_select_actividades()")
        actividades = cursor.fetchall()

    return JsonResponse({
        "is_admin": bool(request.session.get("usuario_admin")),
        "actividades": [
            {
                "id_actividad": a[0],
                "nombre": a[1],
                "descripcion": a[2],
                "horario": a[3],
            }
            for a in actividades
        ]
    })
    
def editar_entrenador(request, id_entrenador):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Método no permitido"}, status=405)
    try:
        nombres = request.POST.get("nombres")
        apellidoP = request.POST.get("apellidoP")
        apellidoM = request.POST.get("apellidoM")
        descripcion = request.POST.get("descripcion")

        file = request.FILES.get("img")
        nombre_img = None

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT url_imagen
                FROM public."Entrenadores"
                WHERE id_entrenador = %s
            """, [id_entrenador])
            row = cursor.fetchone()
            img_actual = row[0] if row else None

        if file:
            carpeta = os.path.join(settings.MEDIA_ROOT, "fotosEntrenadores")
            os.makedirs(carpeta, exist_ok=True)

            ext = os.path.splitext(file.name)[1]
            base = os.path.splitext(file.name)[0]
            nombre_img = f"{base}_{int(time.time())}{ext}"

            ruta_nueva = os.path.join(carpeta, nombre_img)

            with open(ruta_nueva, "wb+") as destino:
                for chunk in file.chunks():
                    destino.write(chunk)

            if img_actual:
                ruta_old = os.path.join(carpeta, img_actual)
                if os.path.exists(ruta_old):
                    os.remove(ruta_old)

        else:
            nombre_img = img_actual

        with connection.cursor() as cursor:
            cursor.callproc(
                "actualizar_entrenador",
                [id_entrenador, nombres, apellidoP, apellidoM, descripcion, nombre_img]
            )
            result = cursor.fetchone()[0]

        return JsonResponse(result)

    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Error al actualizar: {str(e)}"
        }, status=500)

def eliminar_entrenador(request, id_entrenador):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Método no permitido"}, status=405)

    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT url_imagen FROM public."Entrenadores" WHERE id_entrenador = %s', [id_entrenador])
            row = cursor.fetchone()

        if not row:
            return JsonResponse({"success": False, "message": "Entrenador no existe"}, status=404)

        imagen = row[0] 

        with connection.cursor() as cursor:
            cursor.execute("SELECT eliminar_entrenador(%s);", [id_entrenador])
            result = cursor.fetchone()[0]

        if result.get("success") and imagen:
            ruta_img = os.path.join(settings.MEDIA_ROOT, "fotosEntrenadores", imagen)
            if os.path.exists(ruta_img):
                os.remove(ruta_img)

        return JsonResponse(result)

    except Exception as e:
        return JsonResponse({
            "success": False,
            "message": f"Error al eliminar: {str(e)}"
        }, status=500)
        
def agregar_entrenador(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método inválido"}, status=400)

    try:
        nombres = request.POST.get("nombres")
        apellidoP = request.POST.get("apellidoP")
        apellidoM = request.POST.get("apellidoM")
        descripcion = request.POST.get("descripcion")
        file = request.FILES.get("img")

        nombre_img = None

        if file:
            carpeta = os.path.join(settings.MEDIA_ROOT, "fotosEntrenadores")
            os.makedirs(carpeta, exist_ok=True)
            extension = os.path.splitext(file.name)[1]
            nombre_img = f"{uuid.uuid4().hex}{extension}"
            ruta = os.path.join(carpeta, nombre_img)
            with open(ruta, "wb+") as destino:
                for chunk in file.chunks():
                    destino.write(chunk)
                    
        with connection.cursor() as cursor:
            cursor.callproc(
                "insertar_entrenador",
                [nombres, apellidoM, apellidoP, descripcion, nombre_img]
            )

        return JsonResponse({"ok": True})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

def lista_entrenadores_json(request):
    with connection.cursor() as cursor:
        cursor.execute('SELECT * FROM obtener_entrenadores()')
        rows = cursor.fetchall()

    entrenadores_list = []
    for r in rows:
        entrenadores_list.append({
            "id_entrenador": r[0],
            "Nombres": r[1],
            "ApellidoM": r[2],
            "ApellidoP": r[3],
            "descripcion": r[4],
            "url_img": r[5]
        })

    return JsonResponse({
        "entrenadores": entrenadores_list,
        "MEDIA_URL": settings.MEDIA_URL,
        "usuario_admin": request.session.get("usuario_admin", False)
    })
def buscar_usuario_membresia(request):
    if request.method == "POST" and request.headers.get("X-Requested-With") == "XMLHttpRequest":

        try:
            data = json.loads(request.body)
            id_usuario = data.get("usuario")

            if not id_usuario:
                return JsonResponse({
                    "success": False,
                    "error": "ID de usuario requerido"
                }, status=400)

            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT * FROM obtener_membresia_usuario(%s)",
                    [id_usuario]
                )

                columnas = [col[0] for col in cursor.description]
                filas = cursor.fetchall()  
            if not filas:
                return JsonResponse({
                    "success": True,
                    "tiene_membresia": False
                })

            data_sql = dict(zip(columnas, filas[0]))
            print(data_sql)
            return JsonResponse({
                "success": True,
                "tiene_membresia": True,
                "membresia": {
                    "id_membresia": data_sql["id_membresia"],
                    "no_membresia": data_sql["no_membresia"],
                    "nombre_membresia": data_sql["nombre_membresia"],
                    "fecha_inicial": data_sql["fecha_inicial"].strftime("%d/%m/%Y"),
                    "fecha_final": data_sql["fecha_final"].strftime("%d/%m/%Y"),
                    "status": data_sql["status"],
                    "comentario": data_sql["comentario"]
                }
            })

        except Exception as e:
            error = str(e).split("CONTEXT:")[0].strip()
            return JsonResponse({
                "success": False,
                "error": error
            }, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)

def asignar_membresia_usuario_view(request):
    if request.method == "POST" and request.headers.get("X-Requested-With") == "XMLHttpRequest":
        try:
            data = json.loads(request.body)

            id_usuario = data.get("id_usuario")
            id_membresia = data.get("id_membresia")
            fecha_inicio = data.get("fecha_inicio")
            fecha_fin = data.get("fecha_fin")

            if not all([id_usuario, id_membresia, fecha_inicio, fecha_fin]):
                return JsonResponse({
                    "success": False,
                    "error": "Datos incompletos"
                }, status=400)

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT asignar_membresia_usuario(%s, %s, %s, %s)
                """, [
                    id_usuario,
                    id_membresia,
                    fecha_inicio,
                    fecha_fin
                ])

                mensaje = cursor.fetchone()[0]

            return JsonResponse({
                "success": True,
                "mensaje": mensaje
            })

        except Exception as e:
            error = str(e).split("CONTEXT:")[0].strip()
            return JsonResponse({
                "success": False,
                "error": error
            }, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)

def actualizar_membresia(request):
    if request.method == "POST" and request.headers.get("X-Requested-With") == "XMLHttpRequest":
        try:
            data = json.loads(request.body)

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT actualizar_membresia_usuario(%s,%s,%s,%s,%s,%s)
                """, [
                    data.get("id_usuario"),
                    data.get("status"),
                    data.get("id_membresia"),     
                    data.get("fecha_inicial"),    
                    data.get("fecha_final"),       
                    data.get("comentario")
                ])

                mensaje = cursor.fetchone()[0]

            return JsonResponse({
                "success": True,
                "mensaje": mensaje
            })

        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e).split("CONTEXT:")[0]
            }, status=500)

    return JsonResponse({"error": "Método no permitido"}, status=405)

def contador_gimnasio(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT public.contador_gimnasio_hoy();")
        total = cursor.fetchone()[0]

    return JsonResponse({"total": total})   

def agregar_seccion(request):
    if request.method == "POST":
        tipo = request.POST.get("tipo")
        descripcion = request.POST.get("descripcion")

        if not tipo or not descripcion:
            return JsonResponse({
                "success": False,
                "mensaje": "Datos incompletos"
            })

        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT insertar_regla_header(%s, %s);",
                    [tipo, descripcion]
                )
                mensaje = cursor.fetchone()[0]

            if mensaje != "OK":
                return JsonResponse({
                    "success": False,
                    "mensaje": mensaje
                })
            return JsonResponse({
                "success": True,
                "mensaje": "Seccion agregada correctamente"
            })

        except Exception as e:
            return JsonResponse({
                "success": False,
                "mensaje": str(e)
            })
        
def reglas_json(request):
    with connection.cursor() as cursor:
        cursor.execute("""
            SELECT tipo, descripcion
            FROM reglas_header
        """)
        headers = cursor.fetchall()

        cursor.execute("""
            SELECT tipo, regla
            FROM reglas_item
        """)
        items = cursor.fetchall()

    reglas = []
    for tipo, descripcion in headers:
        reglas.append({
            "tipo": tipo,
            "descripcion": descripcion,
            "items": [r for t, r in items if t == tipo]
        })

    return JsonResponse({"reglas": reglas})

def eliminar_seccion(request):
    try:
        data = json.loads(request.body)
        tipo = data.get("tipo")

        if not tipo:
            return JsonResponse({
                "success": False,
                "mensaje": "Sección inválida"
            })
        if not request.session.get("usuario_admin"):
            return JsonResponse({
                "success": False,
                "mensaje": "No autorizado"
            }, status=403)

        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT eliminar_seccion_reglas(%s);
            """, [tipo])

            resultado = cursor.fetchone()[0]

        if resultado != "OK":
            return JsonResponse({
                "success": False,
                "mensaje": resultado
            })

        return JsonResponse({
            "success": True,
            "mensaje": "Sección eliminada correctamente"
        })

    except Exception as e:
        return JsonResponse({
            "success": False,
            "mensaje": f"Error inesperado: {str(e)}"
        })  
        
def guardar_seccion(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)

            tipo = data.get("seccion")      
            reglas = data.get("reglas", [])  

            if not tipo:
                return JsonResponse({
                    "success": False,
                    "mensaje": "Sección inválida"
                })

            with connection.cursor() as cursor:
                cursor.execute("""
                    SELECT guardar_seccion_reglas(%s, %s::jsonb);
                """, [tipo, json.dumps(reglas)])

                resultado = cursor.fetchone()[0]

            if resultado != "OK":
                return JsonResponse({
                    "success": False,
                    "mensaje": resultado
                })

            return JsonResponse({
                "success": True,
                "mensaje": "Sección actualizada correctamente"
            })

        except Exception as e:
            return JsonResponse({
                "success": False,
                "mensaje": f"Error al guardar la sección: {str(e)}"
            })  

def reglas_por_seccion(request, tipo):
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id_regla, regla
                FROM reglas_item
                WHERE tipo = %s
                ORDER BY id_regla
            """, [tipo])

            reglas = [
                {"id": r[0], "regla": r[1]}
                for r in cursor.fetchall()
            ]

        return JsonResponse({
            "success": True,
            "reglas": reglas
        })

    except Exception as e:
        return JsonResponse({
            "success": False,
            "mensaje": str(e)
        }, status=500)
        

def horario_bloque(request, bloque):
    with connection.cursor() as cursor:

        if bloque == "horario":
            cursor.execute("SELECT * FROM obtener_horarios();")
            filas = cursor.fetchall()

            return JsonResponse({
                "datos": [
                    {
                        "id": f[0],
                        "horario": f[1],
                        "usuario": f[2],
                        "costoObs": f[3]
                    } for f in filas
                ]
            })

        elif bloque in ["requisitos_comunidad", "requisitos_equipos"]:
            tipo = "Comunidad ITCJ" if bloque == "requisitos_comunidad" else "Equipos Representativos"

            cursor.execute(
                "SELECT * FROM obtener_requisitos(%s);",
                [tipo]
            )

            return JsonResponse({
                "datos": [r[0] for r in cursor.fetchall()]
            })

        elif bloque == "costos":
            cursor.execute("SELECT * FROM obtener_costos();")
            filas = cursor.fetchall()

            return JsonResponse({
                "datos": [
                    {"tipo_usuario": f[0], "costo": f[1]}
                    for f in filas
                ]
            })

    return JsonResponse({"error": "Bloque no válido"}, status=400)

def horario_bloque_guardar(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "mensaje": "Método inválido"})

    data = json.loads(request.body)
    bloque = data.get("bloque")

    with connection.cursor() as cursor:
        
        if bloque == "horario":
            for h in data["eliminar"]:
                cursor.execute("CALL eliminar_horario(%s);", [h["id"]])

            for h in data["actualizar"]:
                cursor.execute(
                    "CALL actualizar_horario(%s,%s,%s,%s);",
                    [h["id"], h["horario"], h["usuario"], h["costoObs"]]
                )

            for h in data["insertar"]:
                cursor.execute(
                    "CALL insertar_horario(%s,%s,%s);",
                    [h["horario"], h["usuario"], h["costoObs"]]
                )

        elif bloque in ["requisitos_comunidad", "requisitos_equipos"]:
            tipo = "Comunidad ITCJ" if bloque == "requisitos_comunidad" else "Equipos Representativos"

            cursor.execute("CALL eliminar_requisitos_por_tipo(%s);", [tipo])

            for r in data["insertar"]:
                cursor.execute(
                    "CALL insertar_requisito(%s,%s);",
                    [tipo, r["valor"]]
                )
                
        elif bloque == "costos":
            cursor.execute("CALL eliminar_costos();")

            for c in data["insertar"]:
                cursor.execute(
                    "CALL insertar_costo(%s,%s);",
                    [c["tipo_usuario"], c["costo"]]
                )

        else:
            return JsonResponse({"success": False, "mensaje": "Bloque inválido"})

    return JsonResponse({"success": True})

def guardar_noticia(request):
    if request.method == "POST":
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion")
        fecha_publicacion = date.today()

        # 📸 Imagen
        file = request.FILES.get("imagen")
        nombre_img = "noticias/default.png"

        if file:
            carpeta = os.path.join(settings.MEDIA_ROOT, "noticias")
            os.makedirs(carpeta, exist_ok=True)
            extension = os.path.splitext(file.name)[1]
            nombre_img = f"{uuid.uuid4().hex}{extension}"
            ruta = os.path.join(carpeta, nombre_img)

            with open(ruta, "wb+") as destino:
                for chunk in file.chunks():
                    destino.write(chunk)

            nombre_img = f"noticias/{nombre_img}"
        try:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT insertar_noticia(%s, %s, %s, %s);
                    """,
                    [titulo, descripcion, nombre_img, fecha_publicacion]
                )

            return JsonResponse({
                "success": True,
                "mensaje": "Noticia guardada correctamente"
            })

        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            })

    return JsonResponse({
        "success": False,
        "error": "Método no permitido"
    })
    
def listar_noticias(request):
    noticias = []

    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM obtener_noticias();")
            columnas = [col[0] for col in cursor.description]

            for fila in cursor.fetchall():
                noticias.append(dict(zip(columnas, fila)))

        return JsonResponse({
            "success": True,
            "noticias": noticias
        }, safe=False)

    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        }, status=500)
        
def obtener_noticia(request, id):
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT * FROM obtener_noticia_por_id(%s);",
                [id]
            )
            row = cursor.fetchone()

        if not row:
            return JsonResponse({"success": False, "error": "Noticia no encontrada"})

        return JsonResponse({
            "success": True,
            "noticia": {
                "id": row[0],
                "titulo": row[1],
                "descripcion": row[2],
                "imagen": row[3],
            }
        })

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
def actualizar_noticia(request):
    if request.method != "POST":
        return JsonResponse({
            "success": False,
            "error": "Método no permitido"
        })

    try:
        id_noticia = request.POST.get("id_noticia")
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion")

        nueva_imagen = request.FILES.get("imagen")
        ruta_nueva = None

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT imagen FROM noticias WHERE id_noticia = %s;",
                [id_noticia]
            )
            row = cursor.fetchone()

        imagen_actual = row[0] if row else None

        if nueva_imagen:
            carpeta = os.path.join(settings.MEDIA_ROOT, "noticias")
            os.makedirs(carpeta, exist_ok=True)

            extension = os.path.splitext(nueva_imagen.name)[1]
            nombre_img = f"{uuid.uuid4().hex}{extension}"
            ruta_archivo = os.path.join(carpeta, nombre_img)

            with open(ruta_archivo, "wb+") as destino:
                for chunk in nueva_imagen.chunks():
                    destino.write(chunk)

            ruta_nueva = f"noticias/{nombre_img}"

            if imagen_actual:
                ruta_vieja = os.path.join(settings.MEDIA_ROOT, imagen_actual)
                if os.path.exists(ruta_vieja):
                    os.remove(ruta_vieja)

        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT actualizar_noticia(%s, %s, %s, %s);
                """,
                [id_noticia, titulo, descripcion, ruta_nueva]
            )
            success = cursor.fetchone()[0]

        return JsonResponse({"success": success})

    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        })
        
def eliminar_noticia(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"})

    try:
        data = json.loads(request.body)
        id_noticia = data.get("id_noticia")

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT eliminar_noticia_sp(%s);",
                [id_noticia]
            )
            imagen = cursor.fetchone()[0]

        if imagen and imagen != "noticias/default.png":
            ruta_imagen = os.path.join(settings.MEDIA_ROOT, imagen)

            if os.path.exists(ruta_imagen):
                os.remove(ruta_imagen)

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        })
        

def listar_fotos_carrusel(request):
    carpeta = os.path.join(settings.MEDIA_ROOT, "fotosCarrusel")
    imagenes = []

    if os.path.exists(carpeta):
        imagenes = sorted([
            f for f in os.listdir(carpeta)
            if f.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))
        ])

    return JsonResponse({
        "imagenes": [f"fotosCarrusel/{img}" for img in imagenes]
    })
    
def guardar_carrusel(request):
    carpeta_carrusel = os.path.join(settings.MEDIA_ROOT, "fotosCarrusel")
    os.makedirs(carpeta_carrusel, exist_ok=True)

    imagenes_eliminar = json.loads(
        request.POST.get("imagenes_eliminar", "[]")
    )

    for img in imagenes_eliminar:
        ruta = os.path.join(settings.MEDIA_ROOT, img)
        if os.path.exists(ruta):
            os.remove(ruta)

    for nueva_imagen in request.FILES.getlist("imagenes_nuevas"):

        extension = os.path.splitext(nueva_imagen.name)[1]
        nombre_img = f"{uuid.uuid4().hex}{extension}"
        ruta_archivo = os.path.join(carpeta_carrusel, nombre_img)

        with open(ruta_archivo, "wb+") as destino:
            for chunk in nueva_imagen.chunks():
                destino.write(chunk)

    return JsonResponse({"success": True})

REPORTES = {

    "usuarios": {
        "titulo": "Usuarios",
        "descripcion": "El Excel contiene 4 hojas: Alumnos, Empleados, Externos y Representativos.",
        "preview_query": """
            SELECT u.id_usuario, u.nombres, u.apellido_paterno, u.apellido_materno, 'Alumno' AS tipo
            FROM public.alumno a
            JOIN public.usuario u ON u.id_usuario = a.id_usuario
            UNION ALL
            SELECT u.id_usuario, u.nombres, u.apellido_paterno, u.apellido_materno, 'Empleado'
            FROM public.empleado e
            JOIN public.usuario u ON u.id_usuario = e.id_usuario
            UNION ALL
            SELECT u.id_usuario, u.nombres, u.apellido_paterno, u.apellido_materno, 'Externo'
            FROM public.externo ex
            JOIN public.usuario u ON u.id_usuario = ex.id_usuario
            UNION ALL
            SELECT u.id_usuario, u.nombres, u.apellido_paterno, u.apellido_materno, 'Representativo'
            FROM public.representativos r
            JOIN public.usuario u ON u.id_usuario = r.id_usuario
            ORDER BY id_usuario DESC
        """
    },

    "ingresos": {
    "titulo": "Entradas",
    "preview_query": """
        SELECT
            u.id_usuario,
            u.nombres,
            i.fecha,
            i.tipo
        FROM ingresos i
        JOIN public.usuario u ON u.id_usuario = i.id_usuario
          AND (%s IS NULL OR i.fecha >= %s)
          AND (%s IS NULL OR i.fecha <= %s)
        ORDER BY i.fecha DESC
    """
},


"membresias": {
    "titulo": "Membresías",
    "preview_query": """
        SELECT
            u.id_usuario,
            u.nombres,
            m.no_membresia,
            m.fecha_inicial,
            m.fecha_final,
            m.status
        FROM public.membresias m
        JOIN public.usuario u ON u.id_usuario = m.id_usuario
        WHERE (%s IS NULL OR m.fecha_inicial >= %s)
          AND (%s IS NULL OR m.fecha_final <= %s)
        ORDER BY m.fecha_inicial DESC
    """
},


"observaciones": {
    "titulo": "Observaciones",
    "preview_query": """
        SELECT
            id_observacion,
            fecha_observacion,
            titulo,
            descripcion
        FROM public.observaciones
        WHERE (%s IS NULL OR fecha_observacion >= %s)
          AND (%s IS NULL OR fecha_observacion <= %s)
        ORDER BY fecha_observacion DESC
    """
},

}
def preview_reporte(request, tipo):
    if tipo not in REPORTES:
        raise Http404()

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")

    params = [inicio, inicio, fin, fin]

    config = REPORTES[tipo]

    with connection.cursor() as cursor:
        if "%s" in config["preview_query"]:
            cursor.execute(config["preview_query"], params)
        else:
            cursor.execute(config["preview_query"])

        rows = cursor.fetchall()
        headers = [col[0] for col in cursor.description]

    return render(
        request,
        "gym/reportes/preview_modal_content.html",
        {
            "titulo": config["titulo"],
            "headers": headers,
            "rows": rows[:100],
            "total": len(rows),
            "tipo": tipo,
        }
    )
